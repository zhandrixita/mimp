"""
Genera pagina/data/casos_data.js a partir del registro de casos (.sav) del CEM.

Este script se debe volver a ejecutar cada vez que se actualiza la base .sav
(ej. cada mes/corte). Solo escribe los AGREGADOS (conteos y porcentajes) que
necesita la pagina web -- nunca las filas crudas -- para que el archivo JS
que carga el navegador se mantenga pequeno (KB) sin importar cuantos casos
tenga la base (filas).

Salida: pagina/data/casos_data.js
    window.CASOS_DATA = { hombres: {...}, mujeres: {...}, total: {...}, generado: {...} }
"""

import json
from datetime import datetime

import pandas as pd
import pyreadstat
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
from openpyxl.utils import get_column_letter

RUTA_SAV = "data/BD_Registro_casos_julio_2026_SDP.sav"
SALIDA_JS = "pagina/data/casos_data.js"
CARPETA_INFORMES = "pagina/data"

# Color de tema por seccion (mismo hex que --header-green-1 en cada css) y
# titulo de portada -- para que el Excel descargable de cada pestana se vea
# como una extension de esa pestana, no una hoja de calculo generica.
TEMA_INFORME = {
    "hombres": {"color": "1F6D45", "titulo": "Casos de hombres"},
    "mujeres": {"color": "9C2F5C", "titulo": "Casos de mujeres"},
    "alcohol_drogas": {"color": "174F91", "titulo": "Agresor bajo los efectos del alcohol y drogas"},
    "lgtbi": {"color": "6A3FA0", "titulo": "Casos de personas LGBTI"},
    "extranjeras": {"color": "0E8A72", "titulo": "Casos de personas extranjeras"},
    "gestantes": {"color": "B5541F", "titulo": "Mujeres en estado de gestación"},
}

# Historico 2021-2025: no esta en el .sav actual (que solo trae el corte
# preliminar de 2026), son cifras fijas proporcionadas directamente.
# 2026 SI se calcula desde el .sav (ver resumen() / historico_anual).
HISTORICO_ESTATICO = {
    "hombres": {2021: 22964, 2022: 20766, 2023: 24131, 2024: 26348, 2025: 27728},
    "mujeres": {2021: 140833, 2022: 133436, 2023: 142182, 2024: 142144, 2025: 141808},
}
HISTORICO_ESTATICO["total"] = {
    anio: HISTORICO_ESTATICO["hombres"][anio] + HISTORICO_ESTATICO["mujeres"][anio]
    for anio in HISTORICO_ESTATICO["hombres"]
}

COLUMNAS = [
    "SEXO_VICTIMA",
    "EDAD_GRANDE",
    "ESTADO_CIVIL_VICTIMA",
    "DISCAPACIDAD_VICTIMA",
    "VICTIMA_EXTRANJERA",
    "VICTIMA_PERUANA",
    "TRABAJA_VICTIMA",
    "NIVEL_DE_RIESGO_VICTIMA",
    "VINCULO_AGRESOR_VICTIMA",
    "ACOSO_SEX_ESP_PUB",
    "TRATA_CON_FINES_EXPLOTACION_SEXUAL",
    "HOSTIGAMIENTO_SEXUAL",
    "VIOLACION",
    "TIPO_VIOLENCIA",
    "DPTO_UBI_CEM",
    "REGION_UBI_CEM",
    "FECHA_INGRESO",
    # --- TP7: agresor bajo efectos de alcohol y/o drogas ---
    "ESTADO_AGRESOR_U_A",
    "EDAD_AGRESOR",
    "PRIMERA_VEZ_AGREDE",
    # --- TP5: personas LGBTI ---
    "CASOS_PERSONAS_LGBTI",
    "CASOS_PERSONAS_EXTRANJERAS",
    "IDENTIDAD_GENERO",
    "INTERSEXUAL",
    # --- TP6: personas extranjeras ---
    "VICTIMA_GESTANDO",
    "VICTIMA_PAIS_EXTRANJERO",
    # --- mujeres en estado de gestacion ---
    "VICTIMA_TIEMPO_GESTACION",
    "VULNERABILIDAD_VICTIMA_EMBARAZO_VIOLACION",
    # --- v2: perfil de la persona agresora ---
    "SEXO_AGRESOR",
    "EDAD_GRANDE_AGRESOR",
    "NIVEL_EDUCATIVO_AGRESOR",
    "TRABAJA_AGRESOR",
    "DISCAPACIDAD_AGRESOR",
    # --- v2: detalle victima ---
    "NIVEL_EDUCATIVO_VICTIMA",
    "VICTIMA_DISCAPACIDAD_FISICA",
    "VICTIMA_DISCAPACIDAD_VISUAL",
    "VICTIMA_DISCAPACIDAD_AUDITIVA",
    "VICTIMA_DISCAPACIDAD_PSICOSOCIAL",
    "VICTIMA_DISCAPACIDAD_INTELECTUAL",
    "SIS_SEGURO",
    "ESSALUD_SEGURO",
    "PRIVADO_SEGURO",
    "PNP_SEGURO",
    "NINGUN_SEGURO",
    "ETNIA_VICTIMA",
    # --- v2: lugar / ambito ---
    "LUGAR_OCURRENCIA",
    "AMBITO_VIOLENCIA",
    # --- v2: atencion y seguimiento del CEM ---
    "INTERPUSO_DENUNCIA",
    "CUENTA_MEDIDAS_PROTECCION",
    "CUENTA_MEDIDAS_CAUTELARES",
    "ATENCION_INTEGRAL",
    "ATENCION_INTERDISCIPLINARIA",
    "SENTENCIA_FAVORABLE",
    # --- v2: factores de riesgo (persona usuaria) ---
    "FACTOR_VICTIMA_CARENCIA_RED_FAMILIAR",
    "FACTOR_VICTIMA_DEPENDE_ECONOMICAMENTE_AGRESOR",
    "FACTOR_VICTIMA_JUSTIFICA_AGRESIONES",
    "FACTOR_VICTIMA_INTENTA_RETIRAR_DENUNCIA",
    "FACTOR_VICTIMA_INICIA_NUEVA_RELACION",
    "FACTOR_VICTIMA_AISLAMIENTO",
    "FACTOR_VICTIMA_VULNERABILIDAD",
    "FACTOR_VICTIMA_DISCAPACIDAD",
    "FACTOR_VICTIMA_DEPENDE_EMOCIONALMENTE_AGRESOR",
    "FACTOR_VICTIMA_PERCIBE_PELIGRO_DE_MUERTE",
    "FACTOR_VICTIMA_ABUSO_CONSUMO_ALCOHOL",
    "FACTOR_VICTIMA_CONSUME_DROGAS",
    "FACTOR_VICTIMA_HISTORIAL_VIOLENCIA_OTRA_PAREJA",
    "FACTOR_VICTIMA_INDEFENSION",
    "FACTOR_VICTIMA_TENTATIVA_DE_FEMINICIDIO",
    "FACTOR_VICTIMA_PROBLEMA_COMPORTAMENTAL",
    "FACTOR_VICTIMA_INTENTO_DE_SUICIDIO",
    "FACTOR_VICTIMA_INSEGURIDAD_EN_VIVIENDA",
    "FACTOR_VICTIMA_AUSENCIA_DE_CUIDADOR",
    "FACTOR_VICTIMA_OTRO",
    # --- v2: factores de riesgo (presunta persona agresora) ---
    "FACTOR_AGRESOR_VFIS_CAUSA_LESION",
    "FACTOR_AGRESOR_VFIS_PRESENCIA_HIJOS_FAMILIARES",
    "FACTOR_AGRESOR_AMENAZA_CON_OBJETO_PELIGROSO",
    "FACTOR_AGRESOR_ACCESO_ARMA_DE_FUEGO",
    "FACTOR_AGRESOR_AMENAZA_DE_MUERTE",
    "FACTOR_AGRESOR_TIENE_ACCESO_A_VICTIMA",
    "FACTOR_AGRESOR_AUMENTA_EPISODIO_VIOLENTO",
    "FACTOR_AGRESOR_INTENCION_DE_CAUSAR_LESION",
    "FACTOR_AGRESOR_TENTATIVA_DE_FEMINICIDIO",
    "FACTOR_AGRESOR_AGRESION_SEXUAL_DE_PAREJA",
    "FACTOR_AGRESOR_VIOLENTA_HIJOS_FAMILIARES",
    "FACTOR_AGRESOR_INCUMPLE_MEDIDA_PROTECCION",
    "FACTOR_AGRESOR_CELOS_PATOLOGICOS",
    "FACTOR_AGRESOR_HISTORIAL_VIOLENCIA_PAREJA",
    "FACTOR_AGRESOR_HISTORIAL_VIOLENCIA_OTRA_PERSONA",
    "FACTOR_AGRESOR_CONSUMO_ALCOHOL",
    "FACTOR_AGRESOR_CONSUME_DROGA",
    "FACTOR_AGRESOR_ENFERMEDAD_MENTAL",
    "FACTOR_AGRESOR_CONDUCTAS_DE_CRUELDAD",
    "FACTOR_AGRESOR_NEGATIVA_A_SEPARACION",
    "FACTOR_AGRESOR_ANTECEDENTE_LEGAL",
    "FACTOR_AGRESOR_NEGLIGENTE",
    "FACTOR_AGRESOR_LIMITACION_FISICA",
    "FACTOR_AGRESOR_SIN_RED_DE_APOYO",
    "FACTOR_AGRESOR_HISTORIAL_DE_MALTRATO",
    "FACTOR_AGRESOR_RESPUESTA_NEGATIVA",
    "FACTOR_AGRESOR_OTRO",
    # --- v2: sub-actos por tipo de violencia ---
    "PERTURBACION_POSESION",
    "MENOSCABO_TENENCIA_BIENES",
    "PERDIDA_DERECHOS_PATRIMONIALES",
    "LIMITACION_RECURSOS_ECONOMICOS",
    "PRIVACION_MEDIOS_INDISPENSABLES",
    "INCUMPLIMIENTO_OBLIGACION_ALIMENTARIA",
    "CONTROL_DE_INGRESOS",
    "PERCEPCION_SALARIO_MENOR",
    "PROHIBIR_DES_LABORAL",
    "SUSTRAER_INGRESOS",
    "FRACCION_RECURSOS_NEC",
    "OBLIGACION_ALIMENTOS",
    "DESTRUIR_INST_TRABAJO",
    "DESTRUIR_BIEN_PERSONAL",
    "GRITOS_INSULTOS",
    "VIOLENCIA_RACIAL",
    "INDIFERENCIA",
    "DISCR_ORIENTACION_SEXUAL",
    "DISCR_GENERO",
    "DISCR_IDENTIDAD_GENERO",
    "RECHAZO",
    "DESVALORIZACION_HUMILLACION",
    "AMENAZA_QUITAR_HIJOS",
    "OTRAS_AMENAZAS",
    "PROHIBE_RECIBIR_VISITAS",
    "PROHIBE_ESTUDIAR_TRABAJAR_SALIR",
    "ROMPE_DESTRUYE_COSAS",
    "VIGILANCIA_CONTINUA_PERSECUCION",
    "BOTAR_CASA",
    "AMENAZA_DE_MUERTE",
    "ABANDONO",
    "PUNTAPIES_PATADAS",
    "BOFETADAS",
    "JALONES_CABELLO",
    "MORDEDURA",
    "OTRAS_AGRESIONES",
    "EMPUJONES",
    "GOLPES_CON_PALOS",
    "LATIGAZO",
    "AHORCAMIENTO",
    "HERIDAS_CON_ARMAS",
    "GOLPES_CON_OBJETOS_CONTUNDENTES",
    "NEGLIGENCIA",
    "QUEMADURA",
    "EXPLOTACION_SEXUAL",
    "PORNOGRAFIA",
    "EXHIBICION_OBSCENIDAD",
    "PROP_NNA_MED_TEC",
    "ACOSO_SEXUAL",
    "CHANTAJE_SEXUAL",
    "DIF_IMAGEN_CONT_SEX",
    "TOCAMIENTO_SIN_CONSENTIMIENTO",
    "TOCAMIENTO_AGRAVIO_MENORES",
]

# Grupos de sub-actos por tipo de violencia (para el detalle "ver mas").
# Se excluyen un par de columnas cuyo NOMBRE de columna trae un caracter mal
# codificado en el propio .sav (ej. PU\x91ETAZOS) -- no es el bug de
# _des_mojibake (que corrige ETIQUETAS, no nombres de columna), asi que se
# omiten en vez de arriesgar un KeyError.
SUBACTOS_POR_TIPO = {
    "economica": [
        "PERTURBACION_POSESION", "MENOSCABO_TENENCIA_BIENES", "PERDIDA_DERECHOS_PATRIMONIALES",
        "LIMITACION_RECURSOS_ECONOMICOS", "PRIVACION_MEDIOS_INDISPENSABLES",
        "INCUMPLIMIENTO_OBLIGACION_ALIMENTARIA", "CONTROL_DE_INGRESOS", "PERCEPCION_SALARIO_MENOR",
        "PROHIBIR_DES_LABORAL", "SUSTRAER_INGRESOS", "FRACCION_RECURSOS_NEC", "OBLIGACION_ALIMENTOS",
        "DESTRUIR_INST_TRABAJO", "DESTRUIR_BIEN_PERSONAL",
    ],
    "psicologica": [
        "GRITOS_INSULTOS", "VIOLENCIA_RACIAL", "INDIFERENCIA", "DISCR_ORIENTACION_SEXUAL",
        "DISCR_GENERO", "DISCR_IDENTIDAD_GENERO", "RECHAZO", "DESVALORIZACION_HUMILLACION",
        "AMENAZA_QUITAR_HIJOS", "OTRAS_AMENAZAS", "PROHIBE_RECIBIR_VISITAS",
        "PROHIBE_ESTUDIAR_TRABAJAR_SALIR", "ROMPE_DESTRUYE_COSAS", "VIGILANCIA_CONTINUA_PERSECUCION",
        "BOTAR_CASA", "AMENAZA_DE_MUERTE", "ABANDONO",
    ],
    "fisica": [
        "PUNTAPIES_PATADAS", "BOFETADAS", "JALONES_CABELLO", "MORDEDURA", "OTRAS_AGRESIONES",
        "EMPUJONES", "GOLPES_CON_PALOS", "LATIGAZO", "AHORCAMIENTO", "HERIDAS_CON_ARMAS",
        "GOLPES_CON_OBJETOS_CONTUNDENTES", "NEGLIGENCIA", "QUEMADURA",
    ],
    "sexual": [
        "HOSTIGAMIENTO_SEXUAL", "ACOSO_SEX_ESP_PUB", "VIOLACION", "TRATA_CON_FINES_EXPLOTACION_SEXUAL",
        "EXPLOTACION_SEXUAL", "PORNOGRAFIA", "EXHIBICION_OBSCENIDAD", "PROP_NNA_MED_TEC",
        "ACOSO_SEXUAL", "CHANTAJE_SEXUAL", "DIF_IMAGEN_CONT_SEX", "TOCAMIENTO_SIN_CONSENTIMIENTO",
        "TOCAMIENTO_AGRAVIO_MENORES",
    ],
}

FACTORES_VICTIMA = [
    "FACTOR_VICTIMA_CARENCIA_RED_FAMILIAR", "FACTOR_VICTIMA_DEPENDE_ECONOMICAMENTE_AGRESOR",
    "FACTOR_VICTIMA_JUSTIFICA_AGRESIONES", "FACTOR_VICTIMA_INTENTA_RETIRAR_DENUNCIA",
    "FACTOR_VICTIMA_INICIA_NUEVA_RELACION", "FACTOR_VICTIMA_AISLAMIENTO", "FACTOR_VICTIMA_VULNERABILIDAD",
    "FACTOR_VICTIMA_DISCAPACIDAD", "FACTOR_VICTIMA_DEPENDE_EMOCIONALMENTE_AGRESOR",
    "FACTOR_VICTIMA_PERCIBE_PELIGRO_DE_MUERTE", "FACTOR_VICTIMA_ABUSO_CONSUMO_ALCOHOL",
    "FACTOR_VICTIMA_CONSUME_DROGAS", "FACTOR_VICTIMA_HISTORIAL_VIOLENCIA_OTRA_PAREJA",
    "FACTOR_VICTIMA_INDEFENSION", "FACTOR_VICTIMA_TENTATIVA_DE_FEMINICIDIO",
    "FACTOR_VICTIMA_PROBLEMA_COMPORTAMENTAL", "FACTOR_VICTIMA_INTENTO_DE_SUICIDIO",
    "FACTOR_VICTIMA_INSEGURIDAD_EN_VIVIENDA", "FACTOR_VICTIMA_AUSENCIA_DE_CUIDADOR", "FACTOR_VICTIMA_OTRO",
]

FACTORES_AGRESOR = [
    "FACTOR_AGRESOR_VFIS_CAUSA_LESION", "FACTOR_AGRESOR_VFIS_PRESENCIA_HIJOS_FAMILIARES",
    "FACTOR_AGRESOR_AMENAZA_CON_OBJETO_PELIGROSO", "FACTOR_AGRESOR_ACCESO_ARMA_DE_FUEGO",
    "FACTOR_AGRESOR_AMENAZA_DE_MUERTE", "FACTOR_AGRESOR_TIENE_ACCESO_A_VICTIMA",
    "FACTOR_AGRESOR_AUMENTA_EPISODIO_VIOLENTO", "FACTOR_AGRESOR_INTENCION_DE_CAUSAR_LESION",
    "FACTOR_AGRESOR_TENTATIVA_DE_FEMINICIDIO", "FACTOR_AGRESOR_AGRESION_SEXUAL_DE_PAREJA",
    "FACTOR_AGRESOR_VIOLENTA_HIJOS_FAMILIARES", "FACTOR_AGRESOR_INCUMPLE_MEDIDA_PROTECCION",
    "FACTOR_AGRESOR_CELOS_PATOLOGICOS", "FACTOR_AGRESOR_HISTORIAL_VIOLENCIA_PAREJA",
    "FACTOR_AGRESOR_HISTORIAL_VIOLENCIA_OTRA_PERSONA", "FACTOR_AGRESOR_CONSUMO_ALCOHOL",
    "FACTOR_AGRESOR_CONSUME_DROGA", "FACTOR_AGRESOR_ENFERMEDAD_MENTAL",
    "FACTOR_AGRESOR_CONDUCTAS_DE_CRUELDAD", "FACTOR_AGRESOR_NEGATIVA_A_SEPARACION",
    "FACTOR_AGRESOR_ANTECEDENTE_LEGAL", "FACTOR_AGRESOR_NEGLIGENTE", "FACTOR_AGRESOR_LIMITACION_FISICA",
    "FACTOR_AGRESOR_SIN_RED_DE_APOYO", "FACTOR_AGRESOR_HISTORIAL_DE_MALTRATO",
    "FACTOR_AGRESOR_RESPUESTA_NEGATIVA", "FACTOR_AGRESOR_OTRO",
]

# Nivel educativo (12 categorias crudas) consolidado a 5 grupos legibles.
BUCKETS_EDUCACION = {
    1.0: "Sin nivel / Inicial", 2.0: "Sin nivel / Inicial",
    3.0: "Primaria", 4.0: "Primaria",
    5.0: "Secundaria", 6.0: "Secundaria",
    7.0: "Superior", 8.0: "Superior", 9.0: "Superior", 10.0: "Superior",
    11.0: "Básica especial",
    12.0: "Posgrado",
}


def _des_mojibake(texto):
    """
    Revierte un bug de pyreadstat: en una lectura completa (no metadataonly)
    con encoding="latin1", las etiquetas de valor quedan decodificadas dos
    veces (ej. "años" -> "aÃ±os"). En metadataonly no ocurre, lo que confirma
    que es un problema interno de pyreadstat y no del archivo .sav.
    """
    try:
        return texto.encode("latin1").decode("utf-8")
    except (UnicodeEncodeError, UnicodeDecodeError):
        return texto


def _localizar_columna_orientacion_sexual(ruta):
    """
    El NOMBRE de columna 'ORIENTACION_SEXUAL' (no la etiqueta) tambien sufre
    el bug de doble-decodificacion de pyreadstat -- a diferencia de las
    etiquetas de valor/variable, aqui no hay forma de arreglarlo despues
    (usecols debe pasar el nombre EXACTO tal como pyreadstat lo decodifica).
    Se ubica dinamicamente en una pre-lectura de solo metadatos en vez de
    hardcodear el string corrupto en el codigo fuente.
    """
    _, meta_completo = pyreadstat.read_sav(ruta, metadataonly=True, encoding="latin1")
    for columna in meta_completo.column_names:
        mayus = columna.upper()
        if "ORIENTACI" in mayus and "SEXUAL" in mayus and "DISCR" not in mayus and "ESP_OTRO" not in mayus:
            return columna
    return None


def cargar():
    columna_orientacion_cruda = _localizar_columna_orientacion_sexual(RUTA_SAV)
    columnas = COLUMNAS + ([columna_orientacion_cruda] if columna_orientacion_cruda else [])

    # encoding="latin1": el archivo declara UTF-8 en su cabecera pero las
    # etiquetas de valor (ej. "años") estan realmente en Latin-1; sin esto
    # pyreadstat las corrompe (reemplaza los caracteres acentuados por "�").
    df, meta = pyreadstat.read_sav(
        RUTA_SAV, usecols=columnas, apply_value_formats=False, encoding="latin1"
    )

    if columna_orientacion_cruda:
        df.rename(columns={columna_orientacion_cruda: "ORIENTACION_SEXUAL_VICTIMA"}, inplace=True)
        if columna_orientacion_cruda in meta.variable_value_labels:
            meta.variable_value_labels["ORIENTACION_SEXUAL_VICTIMA"] = meta.variable_value_labels.pop(
                columna_orientacion_cruda
            )
        if columna_orientacion_cruda in meta.column_names_to_labels:
            meta.column_names_to_labels["ORIENTACION_SEXUAL_VICTIMA"] = meta.column_names_to_labels.pop(
                columna_orientacion_cruda
            )

    for columna, etiquetas in meta.variable_value_labels.items():
        meta.variable_value_labels[columna] = {
            valor: _des_mojibake(texto) for valor, texto in etiquetas.items()
        }
    # El mismo bug de doble-decodificacion de pyreadstat afecta tambien las
    # etiquetas de VARIABLE (ej. "Factor de riesgo ... : Depende económicamente
    # de la presunta persona agresora"), usadas en top_banderas() para
    # factores de riesgo y sub-actos de violencia.
    meta.column_names_to_labels = {
        col: _des_mojibake(label) for col, label in meta.column_names_to_labels.items()
    }

    # Replica el IF(VICTIMA_EXTRANJERA=1 & VICTIMA_PERUANA=0) del .sps
    df["EXTRANJERO_REPORTE"] = (
        (df["VICTIMA_EXTRANJERA"] == 1) & (df["VICTIMA_PERUANA"] == 0)
    ).astype(int)

    df["ANIO"] = pd.to_datetime(df["FECHA_INGRESO"]).dt.year
    # "YYYY-MM": unico corte de tiempo con datos fila a fila (el .sav solo
    # trae el anio en curso) -- usado por historico_mensual() para el
    # grafico de series de tiempo por sexo/edad de las pestanas tematicas.
    df["MES"] = pd.to_datetime(df["FECHA_INGRESO"]).dt.strftime("%Y-%m")

    return df, meta


def conteo_pct(serie, etiquetas, total):
    """value_counts + % sobre el total, con las etiquetas de valor de SPSS aplicadas."""
    conteo = serie.value_counts(dropna=True)
    salida = {}
    for valor, n in conteo.items():
        clave = etiquetas.get(valor, str(valor)) if etiquetas else str(valor)
        salida[clave] = {"casos": int(n), "pct": round(float(n) / total * 100, 1)}
    return salida


def bandera_pct(serie, total):
    """Cuenta cuantos == 1 en una columna binaria (1=Si, resto=No/perdido)."""
    n = int((serie == 1).sum())
    return {"casos": n, "pct": round(n / total * 100, 1)}


def top_banderas(d, meta, columnas, total, top_n=10):
    """% de cada columna binaria (1=Si) de una lista, con la etiqueta corta
    (lo que sigue despues de ':' en la etiqueta de variable de SPSS),
    ordenado de mayor a menor. Para rankings (factores de riesgo, sub-actos)."""
    filas = []
    for col in columnas:
        n = int((d[col] == 1).sum())
        if n == 0:
            continue
        etiqueta_completa = meta.column_names_to_labels.get(col, col)
        etiqueta = etiqueta_completa.split(":", 1)[-1].strip()
        filas.append({"label": etiqueta, "casos": n, "pct": round(n / total * 100, 1)})
    filas.sort(key=lambda f: -f["casos"])
    return filas[:top_n]


def conteo_por_bucket(serie, buckets, orden, total):
    """Como conteo_pct, pero agrupando valores crudos en buckets (ej. 12
    niveles educativos -> 5 grupos legibles), preservando el orden dado."""
    agrupado = serie.map(buckets)
    conteo = agrupado.value_counts(dropna=True)
    salida = {}
    for clave in orden:
        n = int(conteo.get(clave, 0))
        if n == 0:
            continue
        salida[clave] = {"casos": n, "pct": round(n / total * 100, 1)}
    return salida


ORDEN_EDUCACION = ["Sin nivel / Inicial", "Primaria", "Secundaria", "Superior", "Básica especial", "Posgrado"]


def resumen(df, meta, filtro=None, historico_previo=None):
    d = df if filtro is None else df[filtro]
    total = len(d)

    edad = conteo_pct(d["EDAD_GRANDE"], meta.variable_value_labels.get("EDAD_GRANDE"), total)
    estado_civil = conteo_pct(
        d["ESTADO_CIVIL_VICTIMA"], meta.variable_value_labels.get("ESTADO_CIVIL_VICTIMA"), total
    )
    nivel_riesgo = conteo_pct(
        d["NIVEL_DE_RIESGO_VICTIMA"], meta.variable_value_labels.get("NIVEL_DE_RIESGO_VICTIMA"), total
    )
    vinculo_agresor = conteo_pct(
        d["VINCULO_AGRESOR_VICTIMA"], meta.variable_value_labels.get("VINCULO_AGRESOR_VICTIMA"), total
    )
    tipo_violencia = conteo_pct(
        d["TIPO_VIOLENCIA"], meta.variable_value_labels.get("TIPO_VIOLENCIA"), total
    )

    modalidades_sexuales = {
        "acoso_sexual_espacios_publicos": bandera_pct(d["ACOSO_SEX_ESP_PUB"], total),
        "trata_fines_explotacion_sexual": bandera_pct(d["TRATA_CON_FINES_EXPLOTACION_SEXUAL"], total),
        "hostigamiento_sexual": bandera_pct(d["HOSTIGAMIENTO_SEXUAL"], total),
        "violacion": bandera_pct(d["VIOLACION"], total),
    }

    # DPTO_UBI_CEM: 25 departamentos -- para el mapa (coincide con geodata.js: nombdep)
    por_departamento = conteo_pct(d["DPTO_UBI_CEM"], None, total)
    # REGION_UBI_CEM: 26 categorias, separa Lima Metropolitana / Lima Provincia -- para el listado ordenado
    por_region = conteo_pct(d["REGION_UBI_CEM"], None, total)

    historico_anual = {int(a): int(n) for a, n in d.groupby("ANIO").size().sort_index().items()}
    if historico_previo:
        historico_anual = {**historico_previo, **historico_anual}
    historico_anual = dict(sorted(historico_anual.items()))

    # --- v2: perfil de la persona agresora ---
    agresor_sexo = conteo_pct(d["SEXO_AGRESOR"], meta.variable_value_labels.get("SEXO_AGRESOR"), total)
    agresor_edad = conteo_pct(
        d["EDAD_GRANDE_AGRESOR"], meta.variable_value_labels.get("EDAD_GRANDE_AGRESOR"), total
    )
    agresor_educacion = conteo_por_bucket(d["NIVEL_EDUCATIVO_AGRESOR"], BUCKETS_EDUCACION, ORDEN_EDUCACION, total)

    # --- v2: detalle victima ---
    educacion_victima = conteo_por_bucket(d["NIVEL_EDUCATIVO_VICTIMA"], BUCKETS_EDUCACION, ORDEN_EDUCACION, total)
    discapacidad_detalle = {
        "fisica": bandera_pct(d["VICTIMA_DISCAPACIDAD_FISICA"], total),
        "visual": bandera_pct(d["VICTIMA_DISCAPACIDAD_VISUAL"], total),
        "auditiva": bandera_pct(d["VICTIMA_DISCAPACIDAD_AUDITIVA"], total),
        "psicosocial": bandera_pct(d["VICTIMA_DISCAPACIDAD_PSICOSOCIAL"], total),
        "intelectual": bandera_pct(d["VICTIMA_DISCAPACIDAD_INTELECTUAL"], total),
    }
    seguro_medico = {
        "sis": bandera_pct(d["SIS_SEGURO"], total),
        "essalud": bandera_pct(d["ESSALUD_SEGURO"], total),
        "privado": bandera_pct(d["PRIVADO_SEGURO"], total),
        "pnp_ffaa": bandera_pct(d["PNP_SEGURO"], total),
        "ninguno": bandera_pct(d["NINGUN_SEGURO"], total),
    }
    etnia = conteo_pct(d["ETNIA_VICTIMA"], meta.variable_value_labels.get("ETNIA_VICTIMA"), total)

    # --- v2: lugar / ambito de ocurrencia ---
    lugar_ocurrencia = conteo_pct(
        d["LUGAR_OCURRENCIA"], meta.variable_value_labels.get("LUGAR_OCURRENCIA"), total
    )
    ambito_violencia = conteo_pct(
        d["AMBITO_VIOLENCIA"], meta.variable_value_labels.get("AMBITO_VIOLENCIA"), total
    )

    # --- v2: atencion y seguimiento del CEM ---
    atencion_seguimiento = {
        "denuncia_interpuesta": bandera_pct(d["INTERPUSO_DENUNCIA"], total),
        "medidas_proteccion": bandera_pct(d["CUENTA_MEDIDAS_PROTECCION"], total),
        "medidas_cautelares": bandera_pct(d["CUENTA_MEDIDAS_CAUTELARES"], total),
        "atencion_integral": bandera_pct(d["ATENCION_INTEGRAL"], total),
        "atencion_interdisciplinaria": bandera_pct(d["ATENCION_INTERDISCIPLINARIA"], total),
        "sentencia_favorable": bandera_pct(d["SENTENCIA_FAVORABLE"], total),
    }

    # --- v2: rankings (factores de riesgo, sub-actos por tipo de violencia) ---
    factores_riesgo_victima = top_banderas(d, meta, FACTORES_VICTIMA, total, top_n=8)
    factores_riesgo_agresor = top_banderas(d, meta, FACTORES_AGRESOR, total, top_n=8)
    subactos_violencia = {
        tipo: top_banderas(d, meta, columnas, total, top_n=6)
        for tipo, columnas in SUBACTOS_POR_TIPO.items()
    }

    return {
        "total": total,
        "edad": edad,
        "estado_civil": estado_civil,
        "discapacidad": bandera_pct(d["DISCAPACIDAD_VICTIMA"], total),
        "extranjero": bandera_pct(d["EXTRANJERO_REPORTE"], total),
        "trabaja": bandera_pct(d["TRABAJA_VICTIMA"], total),
        "nivel_riesgo": nivel_riesgo,
        "vinculo_agresor": vinculo_agresor,
        "modalidades_sexuales": modalidades_sexuales,
        "tipo_violencia": tipo_violencia,
        "por_departamento": por_departamento,
        "por_region": por_region,
        "historico_anual": historico_anual,
        "agresor_sexo": agresor_sexo,
        "agresor_edad": agresor_edad,
        "agresor_educacion": agresor_educacion,
        "agresor_trabaja": bandera_pct(d["TRABAJA_AGRESOR"], total),
        "agresor_discapacidad": bandera_pct(d["DISCAPACIDAD_AGRESOR"], total),
        "educacion_victima": educacion_victima,
        "discapacidad_detalle": discapacidad_detalle,
        "seguro_medico": seguro_medico,
        "etnia": etnia,
        "lugar_ocurrencia": lugar_ocurrencia,
        "ambito_violencia": ambito_violencia,
        "atencion_seguimiento": atencion_seguimiento,
        "factores_riesgo_victima": factores_riesgo_victima,
        "factores_riesgo_agresor": factores_riesgo_agresor,
        "subactos_violencia": subactos_violencia,
    }


def resumen_departamento(d, meta, total):
    """
    Version reducida de resumen(), para el detalle por departamento (popup
    al hacer clic en el mapa). Sin por_departamento/por_region/historico
    -- no aportan nada ya filtrado a un solo departamento.
    """
    if total == 0:
        return {"total": 0}
    return {
        "total": total,
        "edad": conteo_pct(d["EDAD_GRANDE"], meta.variable_value_labels.get("EDAD_GRANDE"), total),
        "estado_civil": conteo_pct(
            d["ESTADO_CIVIL_VICTIMA"], meta.variable_value_labels.get("ESTADO_CIVIL_VICTIMA"), total
        ),
        "discapacidad": bandera_pct(d["DISCAPACIDAD_VICTIMA"], total),
        "extranjero": bandera_pct(d["EXTRANJERO_REPORTE"], total),
        "trabaja": bandera_pct(d["TRABAJA_VICTIMA"], total),
        "nivel_riesgo": conteo_pct(
            d["NIVEL_DE_RIESGO_VICTIMA"], meta.variable_value_labels.get("NIVEL_DE_RIESGO_VICTIMA"), total
        ),
        "vinculo_agresor": conteo_pct(
            d["VINCULO_AGRESOR_VICTIMA"], meta.variable_value_labels.get("VINCULO_AGRESOR_VICTIMA"), total
        ),
        "tipo_violencia": conteo_pct(
            d["TIPO_VIOLENCIA"], meta.variable_value_labels.get("TIPO_VIOLENCIA"), total
        ),
        "atencion_seguimiento": {
            "denuncia_interpuesta": bandera_pct(d["INTERPUSO_DENUNCIA"], total),
            "medidas_proteccion": bandera_pct(d["CUENTA_MEDIDAS_PROTECCION"], total),
            "atencion_integral": bandera_pct(d["ATENCION_INTEGRAL"], total),
        },
    }


def construir_detalle_departamentos(df, meta):
    """{ hombres: {depto: resumen_departamento(...)}, mujeres: {...}, total: {...} }"""
    departamentos = sorted(df["DPTO_UBI_CEM"].dropna().unique())
    filtros_sexo = {"hombres": df["SEXO_VICTIMA"] == 1, "mujeres": df["SEXO_VICTIMA"] == 0, "total": None}
    detalle = {"hombres": {}, "mujeres": {}, "total": {}}
    for depto in departamentos:
        base = df["DPTO_UBI_CEM"] == depto
        for clave, filtro_sexo in filtros_sexo.items():
            mask = base if filtro_sexo is None else (base & filtro_sexo)
            sub = df[mask]
            detalle[clave][depto] = resumen_departamento(sub, meta, len(sub))
    return detalle


def _matriz_con_totales(df, fila, columna, filas, columnas):
    """Tabla cruzada con categorias vacias y totales, equivalente a CTABLES."""
    cruce = pd.crosstab(df[fila], df[columna], dropna=False).reindex(
        index=[valor for valor, _ in filas],
        columns=[valor for valor, _ in columnas],
        fill_value=0,
    )
    valores = []
    for valor_fila, _ in filas:
        fila_valores = [int(cruce.loc[valor_fila, valor_col]) for valor_col, _ in columnas]
        valores.append(fila_valores + [sum(fila_valores)])
    totales_columna = [sum(fila[i] for fila in valores) for i in range(len(columnas))]
    valores.append(totales_columna + [sum(totales_columna)])
    return {
        "filas": [etiqueta for _, etiqueta in filas] + ["Total"],
        "columnas": [etiqueta for _, etiqueta in columnas] + ["Total"],
        "valores": valores,
    }


def historico_mensual(d, col_sexo=None, etiquetas_sexo=None, col_edad=None, etiquetas_edad=None, orden_edad=None):
    """
    Serie mensual (el .sav solo trae 2026 fila a fila -- no hay historia
    multi-anio a este nivel de detalle) cruzada opcionalmente por sexo y/o
    grupo etareo. Formato "largo": una fila por combinacion (sexo, edad)
    presente en los datos, cada una con sus conteos mes a mes -- el
    frontend agrupa/suma segun el nivel de detalle que el usuario elija
    (total / por sexo / por grupo de edad / por sexo y edad), asi que un
    solo payload sirve para las 4 vistas del toggle.
    """
    meses = sorted(d["MES"].dropna().unique().tolist())

    def serie_mensual(sub):
        conteo = sub.groupby("MES").size()
        return [int(conteo.get(m, 0)) for m in meses]

    filas = []
    if col_sexo and col_edad:
        combinaciones = d[[col_sexo, col_edad]].dropna().drop_duplicates().values.tolist()
        for sexo_val, edad_val in combinaciones:
            sub = d[(d[col_sexo] == sexo_val) & (d[col_edad] == edad_val)]
            filas.append({
                "sexo": (etiquetas_sexo or {}).get(sexo_val, str(sexo_val)),
                "edad": (etiquetas_edad or {}).get(edad_val, str(edad_val)),
                "valores": serie_mensual(sub),
            })
    elif col_edad:
        for edad_val in sorted(d[col_edad].dropna().unique()):
            sub = d[d[col_edad] == edad_val]
            filas.append({
                "sexo": None,
                "edad": (etiquetas_edad or {}).get(edad_val, str(edad_val)),
                "valores": serie_mensual(sub),
            })
    else:
        filas.append({"sexo": None, "edad": None, "valores": serie_mensual(d)})

    if orden_edad:
        rango_edad = {etiqueta: i for i, etiqueta in enumerate(orden_edad)}
        filas.sort(key=lambda f: (f["sexo"] or "", rango_edad.get(f["edad"], 99)))

    return {"meses": meses, "filas": filas}


def resumen_alcohol_drogas(df, meta):
    """Replica el bloque TP7 del SPSS sobre los estados 2, 3 y 4."""
    d = df[df["ESTADO_AGRESOR_U_A"].isin([2, 3, 4])].copy()
    total = len(d)
    d["EDAD_GRANDE_AGRE"] = pd.cut(
        d["EDAD_AGRESOR"],
        bins=[float("-inf"), 17, 59, float("inf")],
        labels=[1, 2, 3],
    ).astype(float).fillna(9)

    etiquetas_estado = meta.variable_value_labels.get("ESTADO_AGRESOR_U_A", {})
    etiquetas_tipo = meta.variable_value_labels.get("TIPO_VIOLENCIA", {})
    estados = [(2.0, etiquetas_estado.get(2.0, "Efectos de alcohol")),
               (3.0, etiquetas_estado.get(3.0, "Efectos de drogas")),
               (4.0, etiquetas_estado.get(4.0, "Ambos (alcohol y drogas)"))]
    tipos = sorted(
        [(valor, etiqueta) for valor, etiqueta in etiquetas_tipo.items()],
        key=lambda item: item[0],
    )

    return {
        "total": total,
        "pct_base": round(total / len(df) * 100, 1),
        "estado": conteo_pct(d["ESTADO_AGRESOR_U_A"], etiquetas_estado, total),
        "sexo_edad": _matriz_con_totales(
            d, "SEXO_AGRESOR", "EDAD_GRANDE_AGRE",
            [(0.0, "Mujer"), (1.0, "Hombre")],
            [(1.0, "0 a 17 años"), (2.0, "18 a 59 años"),
             (3.0, "60 a más años"), (9.0, "Sin información")],
        ),
        "indicadores": {
            "discapacidad": conteo_pct(
                d["DISCAPACIDAD_VICTIMA"], meta.variable_value_labels.get("DISCAPACIDAD_VICTIMA"), total
            ),
            "extranjero": conteo_pct(d["EXTRANJERO_REPORTE"], {0: "No", 1: "Si"}, total),
            "trabaja": conteo_pct(
                d["TRABAJA_VICTIMA"], meta.variable_value_labels.get("TRABAJA_VICTIMA"), total
            ),
            "primera_vez_agrede": conteo_pct(
                d["PRIMERA_VEZ_AGREDE"], meta.variable_value_labels.get("PRIMERA_VEZ_AGREDE"), total
            ),
        },
        "nivel_riesgo": conteo_pct(
            d["NIVEL_DE_RIESGO_VICTIMA"], meta.variable_value_labels.get("NIVEL_DE_RIESGO_VICTIMA"), total
        ),
        "vinculo_agresor": conteo_pct(
            d["VINCULO_AGRESOR_VICTIMA"], meta.variable_value_labels.get("VINCULO_AGRESOR_VICTIMA"), total
        ),
        "por_departamento": conteo_pct(d["DPTO_UBI_CEM"], None, total),
        "por_region": conteo_pct(d["REGION_UBI_CEM"], None, total),
        "estado_tipo_violencia": _matriz_con_totales(
            d, "ESTADO_AGRESOR_U_A", "TIPO_VIOLENCIA", estados, tipos
        ),
        "historico_mensual": historico_mensual(
            d, "SEXO_AGRESOR", {0.0: "Mujer", 1.0: "Hombre"},
            "EDAD_GRANDE_AGRE", {1.0: "0 a 17 años", 2.0: "18 a 59 años", 3.0: "60 a más años", 9.0: "Sin información"},
            orden_edad=["0 a 17 años", "18 a 59 años", "60 a más años", "Sin información"],
        ),
    }


def _resumen_por_sexo(d, total_base, calculador):
    """
    Parte un dataframe ya filtrado (LGTBI, extranjeras, etc.) en Hombres y
    Mujeres, y aplica calculador(sub, total_sexo) a cada mitad -- asi cada
    pestana tematica puede mostrarse en dos secciones (Hombres/Mujeres) en
    vez de una sola vista mixta.
    """
    resultado = {}
    for valor, clave in ((1.0, "hombres"), (0.0, "mujeres")):
        sub = d[d["SEXO_VICTIMA"] == valor]
        total_sexo = len(sub)
        resultado[clave] = {
            "total": total_sexo,
            "pct_del_total": round(total_sexo / total_base * 100, 1) if total_base else 0.0,
            **calculador(sub, total_sexo),
        }
    return resultado


def resumen_lgtbi(df, meta):
    """Replica el bloque TP5 del SPSS sobre personas LGBTI (CASOS_PERSONAS_LGBTI=1)."""
    d = df[df["CASOS_PERSONAS_LGBTI"] == 1]
    total = len(d)

    def campos(sub, total_sexo):
        return {
            "edad": conteo_pct(sub["EDAD_GRANDE"], meta.variable_value_labels.get("EDAD_GRANDE"), total_sexo),
            "indicadores": {
                "discapacidad": conteo_pct(
                    sub["DISCAPACIDAD_VICTIMA"], meta.variable_value_labels.get("DISCAPACIDAD_VICTIMA"), total_sexo
                ),
                "extranjero": conteo_pct(
                    sub["CASOS_PERSONAS_EXTRANJERAS"], meta.variable_value_labels.get("CASOS_PERSONAS_EXTRANJERAS"), total_sexo
                ),
                "trabaja": conteo_pct(
                    sub["TRABAJA_VICTIMA"], meta.variable_value_labels.get("TRABAJA_VICTIMA"), total_sexo
                ),
            },
            "nivel_riesgo": conteo_pct(
                sub["NIVEL_DE_RIESGO_VICTIMA"], meta.variable_value_labels.get("NIVEL_DE_RIESGO_VICTIMA"), total_sexo
            ),
            "orientacion_sexual": conteo_pct(
                sub["ORIENTACION_SEXUAL_VICTIMA"], meta.variable_value_labels.get("ORIENTACION_SEXUAL_VICTIMA"), total_sexo
            ),
            "identidad_genero": conteo_pct(
                sub["IDENTIDAD_GENERO"], meta.variable_value_labels.get("IDENTIDAD_GENERO"), total_sexo
            ),
            "intersexual": conteo_pct(
                sub["INTERSEXUAL"], meta.variable_value_labels.get("INTERSEXUAL"), total_sexo
            ),
            "tipo_violencia": conteo_pct(
                sub["TIPO_VIOLENCIA"], meta.variable_value_labels.get("TIPO_VIOLENCIA"), total_sexo
            ),
        }

    return {
        "total": total,
        "pct_base": round(total / len(df) * 100, 1),
        **_resumen_por_sexo(d, total, campos),
        # Mantenido para el Excel (tabla cruzada de referencia) -- la vista
        # web ya no lo usa, se reemplazo por las 2 secciones Hombres/Mujeres.
        "sexo_edad": _matriz_con_totales(
            d, "SEXO_VICTIMA", "EDAD_GRANDE",
            [(0.0, "Mujer"), (1.0, "Hombre")],
            [(1.0, "0 a 17 años"), (2.0, "18 a 59 años"), (3.0, "60 a más años")],
        ),
        "por_departamento": conteo_pct(d["DPTO_UBI_CEM"], None, total),
        "por_region": conteo_pct(d["REGION_UBI_CEM"], None, total),
        # Sin cruce por sexo/edad: con solo 97 casos en 7 meses, las celdas
        # sexo x edad quedan en 0-9 (puro ruido) -- se limita a la serie
        # total mensual, que si es representativa.
        "historico_mensual": historico_mensual(d),
    }


def resumen_extranjeras(df, meta):
    """Replica el bloque TP6 del SPSS sobre personas extranjeras (CASOS_PERSONAS_EXTRANJERAS=1)."""
    d = df[df["CASOS_PERSONAS_EXTRANJERAS"] == 1]
    total = len(d)

    def campos(sub, total_sexo):
        return {
            "edad": conteo_pct(sub["EDAD_GRANDE"], meta.variable_value_labels.get("EDAD_GRANDE"), total_sexo),
            "indicadores": {
                "discapacidad": conteo_pct(
                    sub["DISCAPACIDAD_VICTIMA"], meta.variable_value_labels.get("DISCAPACIDAD_VICTIMA"), total_sexo
                ),
                "gestando": conteo_pct(
                    sub["VICTIMA_GESTANDO"], meta.variable_value_labels.get("VICTIMA_GESTANDO"), total_sexo
                ),
                "lgtbi": conteo_pct(
                    sub["CASOS_PERSONAS_LGBTI"], meta.variable_value_labels.get("CASOS_PERSONAS_LGBTI"), total_sexo
                ),
            },
            "nivel_riesgo": conteo_pct(
                sub["NIVEL_DE_RIESGO_VICTIMA"], meta.variable_value_labels.get("NIVEL_DE_RIESGO_VICTIMA"), total_sexo
            ),
            "vinculo_agresor": conteo_pct(
                sub["VINCULO_AGRESOR_VICTIMA"], meta.variable_value_labels.get("VINCULO_AGRESOR_VICTIMA"), total_sexo
            ),
            "pais_origen": conteo_pct(
                sub["VICTIMA_PAIS_EXTRANJERO"], meta.variable_value_labels.get("VICTIMA_PAIS_EXTRANJERO"), total_sexo
            ),
            "tipo_violencia": conteo_pct(
                sub["TIPO_VIOLENCIA"], meta.variable_value_labels.get("TIPO_VIOLENCIA"), total_sexo
            ),
        }

    return {
        "total": total,
        "pct_base": round(total / len(df) * 100, 1),
        **_resumen_por_sexo(d, total, campos),
        # Mantenido para el Excel (tabla cruzada de referencia) -- la vista
        # web ya no lo usa, se reemplazo por las 2 secciones Hombres/Mujeres.
        "sexo_edad": _matriz_con_totales(
            d, "SEXO_VICTIMA", "EDAD_GRANDE",
            [(0.0, "Mujer"), (1.0, "Hombre")],
            [(1.0, "0 a 17 años"), (2.0, "18 a 59 años"), (3.0, "60 a más años")],
        ),
        "por_departamento": conteo_pct(d["DPTO_UBI_CEM"], None, total),
        "por_region": conteo_pct(d["REGION_UBI_CEM"], None, total),
        "historico_mensual": historico_mensual(
            d, "SEXO_VICTIMA", {0.0: "Mujer", 1.0: "Hombre"},
            "EDAD_GRANDE", meta.variable_value_labels.get("EDAD_GRANDE"),
            orden_edad=["0 a 17 años", "18 a 59 años", "60 a más años"],
        ),
    }


ORDEN_TRIMESTRE = [
    "Primer trimestre (1 a 13 semanas)",
    "Segundo trimestre (14 a 27 semanas)",
    "Tercer trimestre (28 a mas semanas)",
]


def resumen_gestantes(df, meta):
    """Mujeres en estado de gestacion (VICTIMA_GESTANDO=1) -- siempre SEXO_VICTIMA=Mujer."""
    d = df[df["VICTIMA_GESTANDO"] == 1].copy()
    total = len(d)

    trimestre = pd.cut(
        d["VICTIMA_TIEMPO_GESTACION"],
        bins=[0, 13, 27, float("inf")],
        labels=ORDEN_TRIMESTRE,
    )

    return {
        "total": total,
        "pct_base": round(total / len(df) * 100, 1),
        "edad": conteo_pct(d["EDAD_GRANDE"], meta.variable_value_labels.get("EDAD_GRANDE"), total),
        "trimestre_gestacion": conteo_pct(trimestre, None, total),
        "indicadores": {
            "discapacidad": conteo_pct(
                d["DISCAPACIDAD_VICTIMA"], meta.variable_value_labels.get("DISCAPACIDAD_VICTIMA"), total
            ),
            "extranjero": conteo_pct(
                d["CASOS_PERSONAS_EXTRANJERAS"], meta.variable_value_labels.get("CASOS_PERSONAS_EXTRANJERAS"), total
            ),
            "lgtbi": conteo_pct(
                d["CASOS_PERSONAS_LGBTI"], meta.variable_value_labels.get("CASOS_PERSONAS_LGBTI"), total
            ),
        },
        "embarazo_por_violacion": bandera_pct(d["VULNERABILIDAD_VICTIMA_EMBARAZO_VIOLACION"], total),
        "nivel_riesgo": conteo_pct(
            d["NIVEL_DE_RIESGO_VICTIMA"], meta.variable_value_labels.get("NIVEL_DE_RIESGO_VICTIMA"), total
        ),
        "vinculo_agresor": conteo_pct(
            d["VINCULO_AGRESOR_VICTIMA"], meta.variable_value_labels.get("VINCULO_AGRESOR_VICTIMA"), total
        ),
        "tipo_violencia": conteo_pct(
            d["TIPO_VIOLENCIA"], meta.variable_value_labels.get("TIPO_VIOLENCIA"), total
        ),
        "por_departamento": conteo_pct(d["DPTO_UBI_CEM"], None, total),
        "por_region": conteo_pct(d["REGION_UBI_CEM"], None, total),
        # Sin cruce por sexo: VICTIMA_GESTANDO=1 implica SEXO_VICTIMA=Mujer
        # siempre, asi que "por sexo" no aportaria una segunda serie.
        "historico_mensual": historico_mensual(
            d, col_edad="EDAD_GRANDE", etiquetas_edad=meta.variable_value_labels.get("EDAD_GRANDE"),
            orden_edad=["0 a 17 años", "18 a 59 años", "60 a más años"],
        ),
    }


def _historico_mensual_filas(historico):
    """Aplana {meses, filas} (formato largo del front) a filas mes/sexo/edad/casos para Excel."""
    filas = []
    for fila in historico["filas"]:
        for mes, casos in zip(historico["meses"], fila["valores"]):
            filas.append({
                "mes": mes, "sexo": fila["sexo"] or "Total", "edad": fila["edad"] or "Total",
                "casos": casos,
            })
    return filas


def _formatear_hoja(hoja, color):
    """
    Encabezado en negrita blanca sobre el color de tema, bordes finos en
    toda la tabla, formato numerico por tipo de columna (casos = miles,
    porcentaje = %), fila de encabezado congelada, autofiltro y ancho de
    columna segun contenido -- para que cada hoja se vea como una tabla de
    informe, no una descarga cruda de datos.
    """
    if hoja.max_row < 1:
        return
    relleno = PatternFill("solid", fgColor=color)
    fino = Side(style="thin", color="D9D9D9")
    borde = Border(left=fino, right=fino, top=fino, bottom=fino)
    encabezados = [c.value for c in hoja[1]]
    for celda in hoja[1]:
        celda.font = Font(bold=True, color="FFFFFF")
        celda.fill = relleno
        celda.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
        celda.border = borde
    for fila in hoja.iter_rows(min_row=2):
        for celda in fila:
            celda.border = borde
    for idx, nombre in enumerate(encabezados, start=1):
        columna = get_column_letter(idx)
        nombre_l = (nombre or "").lower()
        if "porcentaje" in nombre_l or nombre_l == "pct":
            formato = '0.0"%"'
        elif "casos" in nombre_l or nombre_l in ("valor", "posicion", "anio"):
            formato = "#,##0"
        else:
            formato = None
        if formato:
            for celda in hoja[columna][1:]:
                celda.number_format = formato
    hoja.freeze_panes = "A2"
    hoja.auto_filter.ref = hoja.dimensions
    for columna in hoja.columns:
        ancho = min(max(len(str(celda.value or "")) for celda in columna) + 2, 45)
        hoja.column_dimensions[columna[0].column_letter].width = ancho


def _portada(writer, clave, total, extra_filas=None):
    """
    Hoja de caratula (siempre la primera): titulo del informe con el color
    de tema, subtitulo con el periodo, y una mini-tabla de cifras clave
    (casos totales + lo que cada seccion quiera resaltar -- % de la base,
    split hombres/mujeres, indicador destacado, etc.).
    """
    tema = TEMA_INFORME[clave]
    hoja = writer.book.create_sheet("Resumen", 0)

    hoja.merge_cells("A1:C1")
    titulo = hoja["A1"]
    titulo.value = tema["titulo"]
    titulo.font = Font(size=16, bold=True, color="FFFFFF")
    titulo.fill = PatternFill("solid", fgColor=tema["color"])
    titulo.alignment = Alignment(vertical="center", indent=1)
    hoja.row_dimensions[1].height = 32

    hoja.merge_cells("A2:C2")
    subtitulo = hoja["A2"]
    subtitulo.value = "Centro Emergencia Mujer y Familia · Enero - Julio 2026 (preliminar)"
    subtitulo.font = Font(italic=True, color="595959")
    subtitulo.alignment = Alignment(indent=1)

    fila = 4
    filas_clave = [("Casos totales", total, "#,##0")] + (extra_filas or [])
    for etiqueta, valor, formato in filas_clave:
        hoja.cell(row=fila, column=1, value=etiqueta).font = Font(bold=True)
        celda_valor = hoja.cell(row=fila, column=2, value=valor)
        celda_valor.number_format = formato
        fila += 1

    fila += 1
    hoja.cell(row=fila, column=1, value="Generado").font = Font(italic=True, size=9, color="808080")
    hoja.cell(row=fila, column=2, value=datetime.now().strftime("%Y-%m-%d %H:%M")).font = Font(
        italic=True, size=9, color="808080"
    )

    hoja.column_dimensions["A"].width = 30
    hoja.column_dimensions["B"].width = 22


def _escribir_informe(clave, hojas, total, extra_portada=None):
    """Escribe pagina/data/informe_<clave>.xlsx: portada + hojas, todas formateadas."""
    ruta = f"{CARPETA_INFORMES}/informe_{clave}.xlsx"
    color = TEMA_INFORME[clave]["color"]
    with pd.ExcelWriter(ruta, engine="openpyxl") as writer:
        for nombre, filas in hojas.items():
            tabla = pd.DataFrame(filas)
            tabla.to_excel(writer, sheet_name=nombre, index=False)
            _formatear_hoja(writer.sheets[nombre], color)
        _portada(writer, clave, total, extra_portada)
    return ruta


def _hojas_perfil(bloque):
    """Hojas de reporte para un perfil (hombres o mujeres) -- una sola vez, sin
    columna "perfil" (redundante: el archivo entero ya es de ese perfil)."""
    categorias = (
        "edad", "estado_civil", "nivel_riesgo", "vinculo_agresor",
        "tipo_violencia", "agresor_sexo", "agresor_edad",
        "agresor_educacion", "educacion_victima", "etnia",
        "lugar_ocurrencia", "ambito_violencia",
    )
    banderas_simples = (
        "discapacidad", "extranjero", "trabaja", "agresor_trabaja",
        "agresor_discapacidad",
    )
    grupos_banderas = (
        "modalidades_sexuales", "discapacidad_detalle", "seguro_medico",
        "atencion_seguimiento",
    )

    categorias_filas = [
        {"indicador": indicador, "categoria": categoria, "casos": valor["casos"], "porcentaje": valor["pct"]}
        for indicador in categorias
        for categoria, valor in bloque.get(indicador, {}).items()
    ]

    banderas_filas = [
        {
            "grupo": "general", "indicador": indicador,
            "casos": bloque.get(indicador, {}).get("casos", 0),
            "porcentaje": bloque.get(indicador, {}).get("pct", 0),
        }
        for indicador in banderas_simples
    ]
    for grupo in grupos_banderas:
        for indicador, valor in bloque.get(grupo, {}).items():
            banderas_filas.append({
                "grupo": grupo, "indicador": indicador,
                "casos": valor["casos"], "porcentaje": valor["pct"],
            })

    historico_filas = [
        {"anio": anio, "casos": casos} for anio, casos in bloque.get("historico_anual", {}).items()
    ]

    ubicaciones_filas = [
        {"nivel": "departamento", "ubicacion": departamento, "casos": valor["casos"], "porcentaje": valor["pct"]}
        for departamento, valor in bloque.get("por_departamento", {}).items()
    ] + [
        {"nivel": "region", "ubicacion": region, "casos": valor["casos"], "porcentaje": valor["pct"]}
        for region, valor in bloque.get("por_region", {}).items()
    ]

    rankings_filas = []
    for grupo in ("factores_riesgo_victima", "factores_riesgo_agresor"):
        for posicion, valor in enumerate(bloque.get(grupo, []), start=1):
            rankings_filas.append({"grupo": grupo, "tipo_violencia": "", "posicion": posicion, **valor})
    for tipo, filas in bloque.get("subactos_violencia", {}).items():
        for posicion, valor in enumerate(filas, start=1):
            rankings_filas.append({"grupo": "subactos_violencia", "tipo_violencia": tipo, "posicion": posicion, **valor})

    detalle_filas = [
        {
            "departamento": departamento, "indicador": indicador, "categoria": categoria,
            "casos": valor["casos"], "porcentaje": valor["pct"],
        }
        for departamento, valores in bloque.get("por_departamento_detalle", {}).items()
        for indicador in ("edad", "estado_civil", "nivel_riesgo", "vinculo_agresor", "tipo_violencia")
        for categoria, valor in valores.get(indicador, {}).items()
    ]

    return {
        "Categorias": categorias_filas,
        "Indicadores_si": banderas_filas,
        "Historico_anual": historico_filas,
        "Ubicaciones": ubicaciones_filas,
        "Rankings": rankings_filas,
        "Detalle_departamento": detalle_filas,
    }


def _regiones_filas(por_region):
    return [{"region": region, "casos": valor["casos"], "porcentaje": valor["pct"]} for region, valor in por_region.items()]


def _matriz_a_filas(matriz):
    return [
        {"categoria": fila, **dict(zip(matriz["columnas"], valores))}
        for fila, valores in zip(matriz["filas"], matriz["valores"])
    ]


def exportar_informes(data):
    """
    Genera un .xlsx INDEPENDIENTE por pestana (pagina/data/informe_<tab>.xlsx),
    cada uno con su propia portada y solo sus propias tablas -- en vez de un
    unico libro combinado, para que el boton "Descargar" de cada pestana
    entregue justo lo que esa pestana muestra, nada mas.
    """
    rutas = []

    for perfil in ("hombres", "mujeres"):
        bloque = data[perfil]
        rutas.append(_escribir_informe(perfil, _hojas_perfil(bloque), bloque["total"]))

    alcohol = data["alcohol_drogas"]
    hojas_alcohol = {
        "Estado": [
            {"indicador": categoria, "casos": valor["casos"], "porcentaje": valor["pct"]}
            for categoria, valor in alcohol["estado"].items()
        ],
        "Categorias": [
            {"indicador": indicador, "categoria": categoria, "casos": valor["casos"], "porcentaje": valor["pct"]}
            for indicador, categorias_indicador in {
                **alcohol["indicadores"],
                "nivel_riesgo": alcohol["nivel_riesgo"],
                "vinculo_agresor": alcohol["vinculo_agresor"],
            }.items()
            for categoria, valor in categorias_indicador.items()
        ],
        "Regiones": _regiones_filas(alcohol["por_region"]),
        "Sexo_edad": _matriz_a_filas(alcohol["sexo_edad"]),
        "Estado_tipo_violencia": _matriz_a_filas(alcohol["estado_tipo_violencia"]),
        "Historico_mensual": _historico_mensual_filas(alcohol["historico_mensual"]),
    }
    rutas.append(_escribir_informe(
        "alcohol_drogas", hojas_alcohol, alcohol["total"],
        [("% de la base total", alcohol["pct_base"], '0.0"%"')],
    ))

    lgtbi = data["lgtbi"]
    hojas_lgtbi = {
        "Categorias": [
            {"sexo": sexo, "indicador": indicador, "categoria": categoria, "casos": valor["casos"], "porcentaje": valor["pct"]}
            for sexo in ("hombres", "mujeres")
            for indicador, categorias_indicador in {
                "edad": lgtbi[sexo]["edad"],
                **lgtbi[sexo]["indicadores"],
                "nivel_riesgo": lgtbi[sexo]["nivel_riesgo"],
                "orientacion_sexual": lgtbi[sexo]["orientacion_sexual"],
                "identidad_genero": lgtbi[sexo]["identidad_genero"],
                "intersexual": lgtbi[sexo]["intersexual"],
                "tipo_violencia": lgtbi[sexo]["tipo_violencia"],
            }.items()
            for categoria, valor in categorias_indicador.items()
        ],
        "Regiones": _regiones_filas(lgtbi["por_region"]),
        "Sexo_edad": _matriz_a_filas(lgtbi["sexo_edad"]),
        "Historico_mensual": _historico_mensual_filas(lgtbi["historico_mensual"]),
    }
    rutas.append(_escribir_informe(
        "lgtbi", hojas_lgtbi, lgtbi["total"],
        [
            ("% de la base total", lgtbi["pct_base"], '0.0"%"'),
            ("Hombres", lgtbi["hombres"]["total"], "#,##0"),
            ("Mujeres", lgtbi["mujeres"]["total"], "#,##0"),
        ],
    ))

    extranjeras = data["extranjeras"]
    hojas_extranjeras = {
        "Categorias": [
            {"sexo": sexo, "indicador": indicador, "categoria": categoria, "casos": valor["casos"], "porcentaje": valor["pct"]}
            for sexo in ("hombres", "mujeres")
            for indicador, categorias_indicador in {
                "edad": extranjeras[sexo]["edad"],
                **extranjeras[sexo]["indicadores"],
                "nivel_riesgo": extranjeras[sexo]["nivel_riesgo"],
                "vinculo_agresor": extranjeras[sexo]["vinculo_agresor"],
                "pais_origen": extranjeras[sexo]["pais_origen"],
                "tipo_violencia": extranjeras[sexo]["tipo_violencia"],
            }.items()
            for categoria, valor in categorias_indicador.items()
        ],
        "Regiones": _regiones_filas(extranjeras["por_region"]),
        "Sexo_edad": _matriz_a_filas(extranjeras["sexo_edad"]),
        "Historico_mensual": _historico_mensual_filas(extranjeras["historico_mensual"]),
    }
    rutas.append(_escribir_informe(
        "extranjeras", hojas_extranjeras, extranjeras["total"],
        [
            ("% de la base total", extranjeras["pct_base"], '0.0"%"'),
            ("Hombres", extranjeras["hombres"]["total"], "#,##0"),
            ("Mujeres", extranjeras["mujeres"]["total"], "#,##0"),
        ],
    ))

    gestantes = data["gestantes"]
    hojas_gestantes = {
        "Categorias": [
            {"indicador": indicador, "categoria": categoria, "casos": valor["casos"], "porcentaje": valor["pct"]}
            for indicador, categorias_indicador in {
                "edad": gestantes["edad"],
                "trimestre_gestacion": gestantes["trimestre_gestacion"],
                **gestantes["indicadores"],
                "nivel_riesgo": gestantes["nivel_riesgo"],
                "vinculo_agresor": gestantes["vinculo_agresor"],
                "tipo_violencia": gestantes["tipo_violencia"],
            }.items()
            for categoria, valor in categorias_indicador.items()
        ],
        "Regiones": _regiones_filas(gestantes["por_region"]),
        "Historico_mensual": _historico_mensual_filas(gestantes["historico_mensual"]),
    }
    rutas.append(_escribir_informe(
        "gestantes", hojas_gestantes, gestantes["total"],
        [
            ("% de la base total", gestantes["pct_base"], '0.0"%"'),
            ("Embarazo producto de violación", gestantes["embarazo_por_violacion"]["casos"], "#,##0"),
        ],
    ))

    return rutas


def main():
    df, meta = cargar()

    data = {
        "hombres": resumen(df, meta, df["SEXO_VICTIMA"] == 1, HISTORICO_ESTATICO["hombres"]),
        "mujeres": resumen(df, meta, df["SEXO_VICTIMA"] == 0, HISTORICO_ESTATICO["mujeres"]),
        "alcohol_drogas": resumen_alcohol_drogas(df, meta),
        "lgtbi": resumen_lgtbi(df, meta),
        "extranjeras": resumen_extranjeras(df, meta),
        "gestantes": resumen_gestantes(df, meta),
        "generado": {
            "filas_totales": len(df),
            "fuente": RUTA_SAV,
        },
    }

    detalle_departamentos = construir_detalle_departamentos(df, meta)
    for clave in ("hombres", "mujeres"):
        data[clave]["por_departamento_detalle"] = detalle_departamentos[clave]

    # ensure_ascii=True (default): escapa acentos/enies como \uXXXX para que el
    # archivo sea puro ASCII y no dependa de que el servidor declare charset=utf-8
    # al servir el .js (si no lo declara, el navegador puede decodificar mal).
    with open(SALIDA_JS, "w", encoding="ascii") as f:
        f.write("window.CASOS_DATA = ")
        json.dump(data, f, ensure_ascii=True, indent=2)
        f.write(";\n")

    rutas_informes = exportar_informes(data)

    print(f"OK -> {SALIDA_JS}")
    for ruta in rutas_informes:
        print(f"OK -> {ruta}")
    print(f"Filas totales: {data['generado']['filas_totales']:,}")
    print(f"Hombres: {data['hombres']['total']:,} | Mujeres: {data['mujeres']['total']:,}")
    print(f"LGTBI: {data['lgtbi']['total']:,}")
    print(f"Personas extranjeras: {data['extranjeras']['total']:,}")
    print(f"Mujeres en gestacion: {data['gestantes']['total']:,}")


if __name__ == "__main__":
    main()
